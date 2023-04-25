import openpyxl
import re

# Open the Excel file
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active

# Iterate over the first column cells
for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
    # Get the value of the current cell
    cell = row[0]
    if cell.value is not None:
        # Use regular expressions to extract all the numbers before the percentage sign
        matches_percent = re.findall(r'(\d+(,\d+)*)%', cell.value)
        matches_second = re.findall(r'(\d+(,\d+)*)秒', cell.value)
        nums_list = []
        if matches_percent:
            for match in matches_percent:
                nums = match[0].replace(',', '') + "%%"
                nums_list.append(nums)
        if matches_second:
            for match in matches_second:
                nums_list.append(match[0])
        if nums_list:
            # Concatenate the numbers into a string and add square brackets around them
            nums_str = "[" + ','.join(nums_list) + "]"
            cell.offset(0, 1).value = nums_str
        else:
            # If the cell does not contain a percentage sign or the word "秒", leave the second column blank
            cell.offset(0, 1).value = "[]"

# Save the Excel file
wb.save('test.xlsx')

